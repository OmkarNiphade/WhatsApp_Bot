from flask import Flask, request, Response
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os
import re

app = Flask(__name__)

EXCEL_FILE = r"C:\Users\Dell\Desktop\BOTES\Buisness.xlsx"

# Session data storage (for demo, in-memory)
# Key: user phone number, Value: dict with 'state' and collected fields
user_sessions = {}

# ➕ Create Excel if it doesn't exist
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        # Removed "Business Phone" column here
        ws.append([
            "Owner Name", "Business Name", "Business Type",
            "Address", "Working Hours", "Owner Phone",
            "Description", "Services/Products", "Timestamp"
        ])
        wb.save(EXCEL_FILE)

# 🔽 Save validated data to Excel
def save_to_excel(owner, bname, btype, addr, hours, owner_phone, desc, services):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        owner, bname, btype, addr, hours, owner_phone,
        desc, services, datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])
    wb.save(EXCEL_FILE)

# 📖 Read entries
def read_all_entries(limit=5):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]
    rows = rows[-limit:] if limit else rows
    result = ""
    for row in rows:
        # Show owner name, business name, address, timestamp
        result += f"\n• {row[0]} | {row[1]} | {row[3]} | {row[8]}"
    return result.strip() if result else "No data found."

# ❌ Delete entries by name
def remove_by_name(name_to_remove):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data = rows[1:]
    new_data = [row for row in data if str(row[0]).lower() != name_to_remove.lower()]
    if len(data) == len(new_data):
        return False
    else:
        # Remove all sheets and recreate one
        for sheet in wb.sheetnames:
            std = wb[sheet]
            wb.remove(std)
        ws = wb.create_sheet()
        ws.append(headers)
        for row in new_data:
            ws.append(row)
        wb.save(EXCEL_FILE)
        return True

# ✏️ Update owner name
def update_name(old_name, new_name):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    changed = False
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value).lower() == old_name.lower():
            row[0].value = new_name
            changed = True
    wb.save(EXCEL_FILE)
    return changed

# Registration steps order and friendly prompts
# Removed "business_phone" field completely
registration_fields = [
    ("owner", "Please enter the *Owner Name* of the business:"),
    ("business", "Please enter the *Business Name*:"),
    ("type", "Please enter the *Business Type*:"),
    ("address", "Please enter the *Business Address* (village or location):"),
    ("hours", "Please enter the *Working Hours* (e.g., 9am-5pm):"),
    ("owner_phone", "Please enter the *Owner Phone Number* (10 digits):"),
    ("description", "Please provide a short *Description* of the business:"),
    ("services", "Please list the *Services or Products* offered:"),
]

def is_valid_phone(phone):
    return re.fullmatch(r"\d{10}", phone) is not None

@app.route("/whatsapp", methods=["POST"])
def whatsapp_webhook():
    from_number = request.form.get("From", "")
    msg_body = request.form.get("Body", "").strip()
    msg_lower = msg_body.lower()

    # Initialize session if not exists
    if from_number not in user_sessions:
        user_sessions[from_number] = {"state": None, "data": {}}

    session = user_sessions[from_number]

    # === New: Immediate update name command anywhere in conversation ===
    if "->" in msg_body:
        try:
            parts = msg_body.split("->")
            old_name = parts[0].strip()
            new_name = parts[1].strip()
            if update_name(old_name, new_name):
                reply = f"✏️ Name updated from '{old_name}' to '{new_name}'."
            else:
                reply = f"❌ No entry found with the name '{old_name}'."
        except:
            reply = "⚠️ Invalid format for update. Use:\nOldName -> NewName"
        twilio_response = f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Message>{reply}</Message>
</Response>"""
        return Response(twilio_response, mimetype="application/xml")

    # Handle the interactive registration flow
    if session["state"] and session["state"].startswith("register_"):
        current_field = session["state"].replace("register_", "")

        # Phone validation only for owner_phone field now
        if current_field == "owner_phone":
            if not is_valid_phone(msg_body):
                reply = f"❌ Invalid phone number. Please enter a valid 10-digit number for Owner Phone."
                twilio_response = f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Message>{reply}</Message>
</Response>"""
                return Response(twilio_response, mimetype="application/xml")
            else:
                session["data"][current_field] = msg_body
        else:
            session["data"][current_field] = msg_body

        # Find next step
        current_index = next(i for i, f in enumerate(registration_fields) if f[0] == current_field)
        if current_index + 1 < len(registration_fields):
            next_field, prompt = registration_fields[current_index + 1]
            session["state"] = "register_" + next_field
            reply = prompt
        else:
            # All data collected, save to excel
            data = session["data"]
            try:
                save_to_excel(
                    data.get("owner", ""),
                    data.get("business", ""),
                    data.get("type", ""),
                    data.get("address", ""),
                    data.get("hours", ""),
                    data.get("owner_phone", ""),
                    data.get("description", ""),
                    data.get("services", "")
                )
                reply = f"✅ {data.get('owner', '')}, your business '{data.get('business', '')}' has been registered successfully!"
            except Exception as e:
                reply = "⚠️ Failed to save your data. Please try again later."
            # Clear session
            user_sessions[from_number] = {"state": None, "data": {}}

        twilio_response = f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Message>{reply}</Message>
</Response>"""
        return Response(twilio_response, mimetype="application/xml")

    # Entry point: if user says "hi", show menu
    if msg_lower == "hi":
        session["state"] = None
        session["data"] = {}
        reply = (
            "👋 Hello! Welcome to the Business Bot.\n"
            "Please choose an option by replying with the number:\n"
            "1️⃣ Register Business\n"
            "2️⃣ Show Last 5 Entries\n"
            "3️⃣ Show All Entries\n"
            "4️⃣ Search History by Owner Name\n"
            "5️⃣ Remove Entries by Owner Name\n"
            "6️⃣ Update Owner Name\n"
            "\nSend 'help' anytime to see commands."
        )
        twilio_response = f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Message>{reply}</Message>
</Response>"""
        return Response(twilio_response, mimetype="application/xml")

    # Handle menu selection after "hi"
    if session["state"] is None and msg_body in {"1", "2", "3", "4", "5", "6"}:
        option = msg_body
        if option == "1":
            # Start registration flow
            first_field, prompt = registration_fields[0]
            session["state"] = "register_" + first_field
            session["data"] = {}
            reply = "Let's get started with your business registration.\n" + prompt
        elif option == "2":
            reply = "📖 Your last 5 entries:" + read_all_entries(limit=5)
        elif option == "3":
            reply = "📁 All entries:" + read_all_entries(limit=0)
        elif option == "4":
            session["state"] = "awaiting_history_name"
            reply = "Please enter the Owner Name to search history:"
        elif option == "5":
            session["state"] = "awaiting_remove_name"
            reply = "Please enter the Owner Name to remove all entries:"
        elif option == "6":
            session["state"] = "awaiting_update_names"
            reply = "Please send update in the format:\nOldName -> NewName"
        else:
            reply = "Invalid option. Please send a number from 1 to 6."

        twilio_response = f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Message>{reply}</Message>
</Response>"""
        return Response(twilio_response, mimetype="application/xml")

    # Handle awaiting inputs for history, remove, update
    if session["state"] == "awaiting_history_name":
        name = msg_body.strip().lower()
        session["state"] = None
        init_excel()
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))[1:]
        matches = [row for row in data if str(row[0]).lower() == name]
        if matches:
            reply = f"📜 History for '{name.title()}':"
            for row in matches:
                reply += f"\n• {row[0]} | {row[1]} | {row[3]} | {row[8]}"
        else:
            reply = f"❌ No history found for '{name.title()}'"

    elif session["state"] == "awaiting_remove_name":
        name = msg_body.strip()
        session["state"] = None
        if remove_by_name(name):
            reply = f"🗑️ All entries for '{name}' have been deleted."
        else:
            reply = f"❌ No entry found with the name '{name}'."

    elif session["state"] == "awaiting_update_names":
        try:
            parts = msg_body.split("->")
            old_name = parts[0].strip()
            new_name = parts[1].strip()
            session["state"] = None
            if update_name(old_name, new_name):
                reply = f"✏️ Name updated from '{old_name}' to '{new_name}'."
            else:
                reply = f"❌ No entry found with the name '{old_name}'."
        except:
            reply = "⚠️ Invalid format. Use:\nOldName -> NewName"

    elif msg_lower == "help":
        reply = (
            "🤖 *Supported Commands:*\n"
            "• Send 'hi' to start the menu.\n"
            "• Or use commands:\n"
            "Register:\n"
            "Owner=Your Name\n"
            "Business=Business Name\n"
            "Type=Type of Business\n"
            "Address=Village or Location\n"
            "Hours=Timing\n"
            "Owner Phone=10-digit number\n"
            "Description=Brief description\n"
            "Services=Products or Services\n"
            "You can also use menu options after 'hi'."
        )
    else:
        reply = "⚠️ Sorry, I did not understand that. Send 'hi' to start."

    twilio_response = f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
    <Message>{reply}</Message>
</Response>"""
    return Response(twilio_response, mimetype="application/xml")

if __name__ == "__main__":
    app.run(debug=True)
